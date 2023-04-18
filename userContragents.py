import random

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

workbook = openpyxl.Workbook()
worksheet = workbook.active

headers = ['Айди','Город','Номер счета','Наименование','users_id']
for col_num, header in enumerate(headers, 1):
    col_letter = get_column_letter(col_num)
    worksheet[f'{col_letter}1'] = header



ids = []

number = []
name_of_busnes = []
users_idd = []
for i in range(5000):
    ids.append(i)
    

for i in range(5000):
    random_number=random.randint(0, 999)
    users_idd.append(random_number)
#Называние компаний
companies = ['ElevateTech', 'IntegroCore', 'NimbusFlow', 'ZenithWorks', 'InnovateForge', 'VeritasLabs', 'GlobalOptima', 'HorizonX', 'UnityEnterprises', 'QuantumSynergy', 'InnovateTechnologies', 'MomentumLabs', 'AvantGardeCorp', 'PlatinumWorks', 'RevolutionWorks', 'PinnacleInnovations', 'RenaissanceCorp', 'EmpireEnterprises', 'AdvancedTech', 'VentureWorks', 'AscentEnterprises', 'VanguardInnovations', 'PrimeLabs', 'ProvenanceCorp', 'FusionWorks', 'AccelCore', 'SummitEnterprises', 'InfiniteLabs', 'ElevateEnterprises', 'AvalonWorks', 'NexusLabs', 'VelocityInnovations', 'InnovaCorp', 'StrategicEnterprises', 'ElementalLabs', 'EndeavorWorks', 'ExponentialCorp', 'SynchronicityInnovations', 'QuantumWorks', 'ElevateLabs', 'PinnacleEnterprises', 'RapidCore', 'HorizonInnovations', 'DynamicEnterprises', 'ElysiumCorp', 'ProvenanceLabs', 'AscentWorks', 'VentureInnovations', 'InfiniteEnterprises', 'RenaissanceWorks', 'ZenithLabs', 'MomentumCorp', 'AvantGardeInnovations', 'InnovateCore', 'PlatinumEnterprises', 'VeritasEnterprises', 'NimbusLabs', 'EmpireWorks', 'QuantumInnovations', 'RevolutionEnterprises', 'FusionCorp', 'AdvancedEnterprises', 'AccelWorks', 'ProvenanceInnovations', 'SummitInnovations', 'AscentLabs', 'PinnacleCorp', 'VanguardEnterprises', 'PrimeEnterprises', 'InfiniteCore', 'ElementalWorks', 'EndeavorLabs', 'VelocityEnterprises', 'ExponentialWorks', 'StrategicInnovations', 'SynchronicityEnterprises', 'ElysiumEnterprises', 'ElevateInnovations', 'PinnacleLabs', 'RapidInnovations', 'HorizonEnterprises', 'DynamicWorks', 'ZenithEnterprises', 'RenaissanceInnovations', 'InnovaLabs', 'MomentumEnterprises', 'AvantGardeWorks', 'VeritasInnovations', 'NimbusEnterprises', 'EmpireInnovations', 'QuantumCorp', 'RevolutionLabs', 'FusionEnterprises', 'AdvancedInnovations', 'AccelInnovations', 'SummitWorks', 'AscentCorp', 'PinnacleInnovations', 'VanguardWorks', 'PrimeLabs', 'InfiniteInnovations', 'ElementalEnterprises', 'EndeavorEnterprises', 'VelocityWorks', 'ExponentialEnterprises', 'StrategicWorks', 'SynchronicityLabs', 'ElysiumInnovations']
companies2 = ['ТОО','ИП','КХ','КТ','АО']
for  i in range(5000):
    name = random.choice(companies)
    name2 = random.choice(companies2)
    last_name_of_bus = name2 + '' + name
    name_of_busnes.append(last_name_of_bus)


#Номер счета
for i in range(5000):
    number.append("0" + str(random.randint(0, 9999)).zfill(4))

cities = [
    "Алматы",
    "Алатау",
    "Алмалинский",
    "Арасан",
    "Ауэзовский",
    "Баганашыл",
    "Байсерке",
    "Байтурсын",
    "Ботанический",
    "Жетысай",
    "Илийский",
    "Калкаман",
    "Каскелен",
    "Коктал",
    "Кокшетау",
    "Медеуский",
    "Мерей",
    "Мкр Комсомольский",
    "Мкр Мамыр",
    "Мкр Наурызбай",
    "Мкр Шалдыбан",
    "Молодежный",
    "Наурыз",
    "Первомайский",
    "Райымбека",
    "Сайран",
    "Самал",
    "Сарыагаш",
    "Сарыарка",
    "Сатпаев",
    "Талгарский",
    "Турксибский",
    "Шу",
    "Шымбулак",
    "Шынгыс",
    "Жанаозен",
    "Кульсары",
    "Тараз",
    "Талдыкорган",
    "Текели",
    "Уштобе",
    "Шардара",
    "Есик",
    "Капчагай",
    "Жаркент",
    "Каскеле",
    "Талгар",
    "Шемонаиха"
]
city = []
for i in range(5000):
    city1 = random.choice(cities)
    city.append(city1)



#Создадим таблицу
data = {
    'Айди': ids,
    "Город":city
    ,
    "Номер счета" :number,
    "Наименрвание":name_of_busnes,
    "users_id":users_idd
}
df = pd.DataFrame(data)

for row_num, data in enumerate(zip(ids,city, number, name_of_busnes, users_idd), 2):
    for col_num, cell_data in enumerate(data, 1):
        col_letter = openpyxl.utils.get_column_letter(col_num)
        worksheet[f'{col_letter}{row_num}'] = cell_data

workbook.save('userContragents_last.xlsx')