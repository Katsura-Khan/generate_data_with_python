import random

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

workbook = openpyxl.Workbook()
worksheet = workbook.active

headers = ['id','usercontragent_id','артикул товара ','бренд товара ','количества товара']
for col_num, header in enumerate(headers, 1):
    col_letter = get_column_letter(col_num)
    worksheet[f'{col_letter}1'] = header


usercontragent_id = []
arti_tov = []
brand = []
count_tov = []
idd = []
brands = ['Adidas', 'Apple', 'Amazon', 'Acer', 'Asus', 'Bosch', 'Canon', 'Coca-Cola', 'Calvin Klein', 'Dell', 
          'Diesel', 'Dyson', 'Dior', 'Epson', 'Estée Lauder', 'Ferrari', 'Ford', 'Fujifilm', 'Gucci', 'Google', 
          'GoPro', 'Harley-Davidson', 'Honda', 'HP', 'H&M', 'IBM', 'Intel', 'JBL', 'Kia', 'Lacoste', 'Lenovo', 
          'LG', 'Logitech', 'Lamborghini', 'Levi\'s', 'Lego', 'Loreal', 'Maserati', 'Mercedes-Benz', 'Microsoft', 
          'Mitsubishi', 'Motorola', 'Nestle', 'Nike', 'Nikon', 'Nivea', 'Nokia', 'Oakley', 'Panasonic', 'Pepsi', 
          'Philips', 'Puma', 'Porsche', 'Prada', 'Ralph Lauren', 'Ray-Ban', 'Red Bull', 'Reebok', 'Samsung', 'Sony', 
          'Siemens', 'Skoda', 'Swarovski', 'Suzuki', 'Tiffany & Co.', 'Tommy Hilfiger', 'Tesla', 'Toyota', 'Toshiba', 
          'Uber', 'Under Armour', 'Vans', 'Versace', 'Volkswagen', 'Volvo', 'Xerox', 'Xiaomi', 'Yamaha', 'Zara', 
          'Zegna', 'Zara Home', 'Adidas Originals', 'Armani', 'Burberry', 'Cartier', 'Chanel', 'Christian Dior', 
          'Converse', 'Dolce & Gabbana', 'Fendi', 'Givenchy', 'Gucci Beauty', 'Hermes', 'Louis Vuitton', 'New Balance', 
          'Prada Group', 'Puma SE', 'Rolex', 'Salvatore Ferragamo', 'Supreme', 'The North Face', 'Tiffany & Co.', 
          'Tod\'s', 'Tom Ford', 'Vans Shoes', 'Yves Saint Laurent']

for i in range(1000000):
    idd.append(i)

for i in range(1000000):
    qw = random.randint(1,10)
    count_tov.append(qw)

for i in range(1000000):
    cv=random.randint(1,999)
    usercontragent_id.append(cv)

for i in range(1000000):
    arti_tov.append("0" + str(random.randint(0, 9999)).zfill(4))

for i in range(1000000):
    er = random.choice(brands)
    brand.append(er)

data = {
    "id":idd,
    "usercontragent_id":usercontragent_id,
    "Артикул товара":arti_tov,
    "Бренд товара":brand,
    "Количество товора":count_tov
}
df = pd.DataFrame(data)

for row_num, data in enumerate(zip(idd,usercontragent_id, arti_tov, brand, count_tov), 2):
    for col_num, cell_data in enumerate(data, 1):
        col_letter = openpyxl.utils.get_column_letter(col_num)
        worksheet[f'{col_letter}{row_num}'] = cell_data

workbook.save('order_history.xlsx')