import random
import string
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

workbook = openpyxl.Workbook()
worksheet = workbook.active


headers = ['ID','ФИО', 'Логин', 'Пароль', 'Тип лица']
for col_num, header in enumerate(headers, 1):
    col_letter = get_column_letter(col_num)
    worksheet[f'{col_letter}1'] = header
# Список ФИО
full_names = []
idd = []
for i in range(1000):
    idd.append(i)



# Список возможных имен и фамилий
possible_names = ['Александр', 'Иван', 'Мария', 'Елена', 'Дмитрий', 'Анна', 'Николай', 'Ольга', 'Андрей', 'Анастасия']
possible_surnames = ['Иванов', 'Петров', 'Сидоров', 'Кузнецова', 'Смирнов', 'Михайлова', 'Федоров', 'Новиков', 'Алексеева', 'Козлов']

# Создание списка из 1000 случайных ФИО
for i in range(1000):
    name = random.choice(possible_names)
    surname = random.choice(possible_surnames)
    full_name = surname + ' ' + name
    full_names.append(full_name)

# Список логинов и паролей
logins = []
passwords = []
email = ['@gmail.com','@mail.ru','@bk.ru','@egov.kz','@hotmail.com']

# Создание списка логинов и паролей
for i in range(1000):
    # Генерация случайного логина
    letters = string.ascii_lowercase
    login = ''.join(random.choice(letters) for i in range(8))
    random_email = random.choice(email)
    last_login = login + random_email
    logins.append(last_login)

    # Генерация случайного пароля
    letters_digits = string.ascii_letters + string.digits
    password = ''.join(random.choice(letters_digits) for i in range(10))
    passwords.append(password)

# Список типов лиц
types = []

# Создание списка типов лиц
for i in range(1000):
    # Генерация случайного типа лица
    type_person = random.choice(['юр лицо', 'физ лицо'])
    types.append(type_person)

# Создание таблицы
data = {'Айди':idd,
        'ФИО': full_names,
        'Логин': logins,
        'Пароль': passwords,
        'Тип лица': types}

df = pd.DataFrame(data)

for row_num, data in enumerate(zip(idd,full_names, logins, passwords, types), 2):
    for col_num, cell_data in enumerate(data, 1):
        col_letter = openpyxl.utils.get_column_letter(col_num)
        worksheet[f'{col_letter}{row_num}'] = cell_data



workbook.save('data4.xlsx')





